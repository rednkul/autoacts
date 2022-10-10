import sys, os, shutil
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.cell.cell import MergedCell
import datetime

from django.conf import settings


# Порядок работы функций:
# 1. create_acts_files
# 2. make_dir
# 3. act_from_template
# 4. copy_template_to_project_dir
# 5. get_act_name
# 6. excel_processing
# Список клеток, которые нужно разъединить:
# works;
# materials
# даты начала-конца


def create_acts_files(project):
    acts = project.acts.all()
    template = project.template
    make_dir(project)

    for act in acts:
        act_from_template(act, template)
    os.chdir(settings.BASE_DIR)


def copy_template_to_project_dir(template):
    template_url = settings.MEDIA_ROOT + '\\' + str(template).replace('/', '\\')  # Copy function requires an ABS path
    return shutil.copy(template_url, os.getcwd())


def make_dir(project):
    """Make project directory which will includes generated acts"""
    os.chdir("media/projects/excel")
    project_dir_name = f'{project.name}_{project.id}'
    if not os.path.isdir(project_dir_name):
        os.mkdir(project_dir_name)
    os.chdir(project_dir_name)


def get_act_name(act):
    return f'{act.order_number}. {act.act_type.name}.xlsx'


def act_from_template(act, template):
    temp_file = os.path.basename(rf'{copy_template_to_project_dir(template)}')
    excel_processing(act, temp_file)


def excel_processing(act, temp_file):
    excel_file = openpyxl.load_workbook(temp_file)
    sheet = excel_file["Лист 1"]  # excel_sheet

    # По опорному слову "Акт" получаю ячейку номера акта и вписываю значение
    number_cell_coordinate = find_act_number_cell(sheet)
    number_cell = sheet[f'{number_cell_coordinate}']
    number_cell.value = act.number

    # По ячейке номера акта получаю ячейку даты акта и вписываю значение
    act_date_cell = sheet.cell(row=number_cell.row, column=number_cell.column + 5)
    act_date_cell.value = act.date_end

    # Работы
    works_cell_row = insert_field_and_return_row_number(sheet, string=act.works, start_row=act_date_cell.row, field=1)

    # Материалы
    materials_string = ', '.join([name for name in act.materials.all().values_list('name', flat=True)])
    materials_cell_row = insert_field_and_return_row_number(sheet, materials_string, start_row=works_cell_row, field=3)

    # Даты начала - конца
    dates_cell_coordinate = find_cell_coordinate_by_value(
        sheet,
        '5.  Даты:',
        start_row=materials_cell_row,

    )
    dates_cell = sheet[f'{dates_cell_coordinate}']

    date_start_cell = sheet.cell(dates_cell.row, dates_cell.column)

    date_start_cell.value = act.date_start
    date_end_cell = sheet.cell(date_start_cell.row + 1, date_start_cell.column)
    date_end_cell.value = act.date_end

    # Следующие работы
    if act.project.acts.filter(order_number=act.order_number + 1).exists():
        next_works = act.project.acts.get(order_number=act.order_number + 1).works
        next_works_cell_row = insert_field_and_return_row_number(sheet, string=next_works, start_row=dates_cell.row,
                                                                 field=7)

    # Количество экземпляров акта
    number_of_instances_cell_coordinate = find_cell_coordinate_by_value(
        sheet,
        'Количество экземпляров акта, шт. :',
        start_row=date_end_cell.row,
    )

    number_of_instances_cell = sheet[f'{number_of_instances_cell_coordinate}']
    number_of_instances_cell.value = act.number_of_instances

    # excel_file.save(get_act_name(act))
    # excel_file = openpyxl.load_workbook(get_act_name(act))
    # sheet = excel_file["Лист 1"]
    # works_cell = find_works_cell(number_cell.row, sheet)
    # prepare_cells(sheet, works_cell.row + 1, works_cell.row + len(works_string_list) - 1)

    excel_file.save(get_act_name(act))


def insert_field_and_return_row_number(sheet, string: str, start_row: int, field: int, row_range=100):
    """
        Функция принимает лист, вставляемую строку, начальную строку листа и поле( работы, материалы,
        следующие работы, приложения). Возвращает номер найденной строки, который понадобится для поиска
        следующей опорной строки
    """

    if field == 1:
        lookup_string = '1. К освидетельствованию предъявлены следующие работы'
    elif field == 3:
        lookup_string = '3. При выполнении работ применены'
    elif field == 7:
        lookup_string = '7. Разрешается  производство   последующих  работ  по'
    # По строке номера получаю строку работ
    cell_coordinate = find_cell_coordinate_by_value(
        sheet,
        lookup_string,
        start_row=start_row,
        row_range=row_range,
    )
    cell = sheet[f'{cell_coordinate}']
    # Поле works делю на отдельные строки с фиксированной длинной и получаю список строк
    string_list = prepare_string_list(string)

    # Вставляю новые строки в зависимости от длины списка строк
    insert_rows(sheet, cell.row + 1, len(string_list))
    insert_string_list(sheet, cell.row, string_list)
    return cell.row


def prepare_cells(sheet, start_row, end_row):
    for row in range(start_row, end_row):
        for merged_range in sheet.merged_cells.ranges:

            if sheet.cell(row=row, column=6).coordinate in merged_range:
                print(merged_range)
                print(sheet.cell(row=row, column=6).coordinate)
                print(f'F{row} merged')
                sheet.unmerge_cells(f'{merged_range}')

        sheet.row_dimensions[row].height = 15.75


def insert_rows(sheet, row, number):
    if number > 1:
        for i in range(number - 1):
            sheet.insert_rows(row)


def insert_string_list(sheet, start_row, string_list):
    sheet[f'F{start_row}'].value = string_list[0]
    list_len = len(string_list)
    if list_len > 1:
        for i, row in enumerate(range(start_row + 1, start_row + len(string_list))):
            sheet[f'A{row}'] = string_list[i + 1]


def find_act_number_cell(sheet):
    for row in list(sheet.rows)[:300]:

        if row[0].value == 'Акт':
            return sheet.cell(row=row[0].row + 2, column=2).coordinate

    else:
        raise NotFoundReferencedWord('Акт')


def find_cell_coordinate_by_value(sheet, value, column=0, start_row=1, row_range=100, ):
    """
    Принимает лист файла эксельБ значение искомой ячейки, ее колонку, и диапазон строк.
    Возвращает координаты искомой ячейки. Если значение не найдено, поднимает исключение.
    """
    for row in list(sheet.rows)[start_row: start_row + row_range]:
        if row[column].value == value:
            return sheet.cell(row=row[0].row, column=6).coordinate
    else:
        raise NotFoundReferencedWord(f'{value}')


# def find_works_cell(start_row, sheet):
#     for row in list(sheet.rows)[start_row: start_row + 100]:
#         if row[0].value == '1. К освидетельствованию предъявлены следующие работы':
#             return sheet.cell(row=row[0].row, column=7)
#     else:
#         raise NotFoundReferencedWord('Работы')


def prepare_string_list(string):
    """
        Поля работ, материалов и приложений в исполнительном акте могут занимать больше одной строки.
        Функция принимает на вход поле модели или сложенную из списка материалов строку и возвращает список строк,
        каждая из которых имеет длину ячейки, в которую она будет помещена.
    """

    string_list = [string[:50]]
    if len(string) > 50:
        for i in range(len(string[50:]) // 95 + 1):
            string_list.append(string[50 + 95 * i:50 + 95 * (i + 1)])

    return string_list


class NotFoundReferencedWord(Exception):
    """Не найдено опорное слово"""

    def __init__(self, referenced_word: str):
        self.referenced_word = referenced_word
        self.message = f"Не найдено опорное слово: '{self.referenced_word}'."

    def __str__(self):
        return self.message
